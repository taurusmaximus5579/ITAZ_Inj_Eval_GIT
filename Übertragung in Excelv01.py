# -*- coding: utf-8 -*-
"""
Created on Mon Nov 10 15:15:26 2025

@author: TC-Laptop-CAD2
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

root = tk.Tk()
root.title("Datenübertragung Excel (transponiert)")
root.geometry("950x700")

# --- Variablen ---
datei_frames = {}
datei_settings = {}
ziel_var = tk.StringVar()

quellblaetter = ["Mass (mg)", "System Pressure (abs_bar)", "Needle Lift (mm)", "Needle Lift Integrated (mm_s)"]
zielzellen = ["A3", "E3", "I3", "M3"]

druck_optionen = ["20 bar", "50 bar", "70 bar"]
zeit_optionen = ["2 ms", "5 ms", "12,5 ms", "18 ms"]

# --- Funktionen ---
def select_quellen():
    dateien = filedialog.askopenfilenames(title="Quelldateien wählen", filetypes=[("Excel-Dateien", "*.xlsx *.xls")])
    if dateien:
        for pfad in list(dateien)[:12]:
            if pfad not in datei_settings:
                datei_settings[pfad] = {
                    "druck": tk.StringVar(value="20 bar"),
                    "zeit": tk.StringVar(value="2 ms"),
                    "zeilen": [tk.StringVar(value="3") for _ in range(4)]
                }
        update_frames()

def select_ziel():
    datei = filedialog.askopenfilename(title="Zieldatei wählen", filetypes=[("Excel-Dateien", "*.xlsx *.xls")])
    if datei:
        ziel_var.set(datei)

def update_frames():
    for widget in scroll_frame.winfo_children():
        widget.destroy()
    datei_frames.clear()
    
    for pfad, settings in datei_settings.items():
        frame = tk.Frame(scroll_frame, padx=5, pady=5, bd=1, relief="solid")
        frame.pack(fill="x", padx=5, pady=3)

        tk.Label(frame, text=os.path.basename(pfad), width=25, anchor="w").grid(row=0, column=0, padx=2, pady=2)

        tk.Label(frame, text="Druck:").grid(row=0, column=1, padx=2, pady=2)
        tk.OptionMenu(frame, settings["druck"], *druck_optionen).grid(row=0, column=2, padx=2, pady=2)

        tk.Label(frame, text="Zeit:").grid(row=0, column=3, padx=2, pady=2)
        tk.OptionMenu(frame, settings["zeit"], *zeit_optionen).grid(row=0, column=4, padx=2, pady=2)

        for i, blatt in enumerate(quellblaetter):
            tk.Label(frame, text=f"Zeile '{blatt}':").grid(row=0, column=5 + i*2, padx=2, pady=2)
            tk.Entry(frame, textvariable=settings["zeilen"][i], width=5).grid(row=0, column=6 + i*2, padx=2, pady=2)

        datei_frames[pfad] = frame

def starten():
    if not datei_settings:
        messagebox.showerror("Fehler", "Keine Quelldateien ausgewählt.")
        return
    ziel_datei = ziel_var.get()
    if not ziel_datei:
        messagebox.showerror("Fehler", "Zieldatei auswählen.")
        return

    verschiebung_mapping = {"2 ms":0, "5 ms":1, "12,5 ms":2, "18 ms":3}

    try:
        wb_ziel = load_workbook(ziel_datei)

        for pfad, settings in datei_settings.items():
            druckwahl = settings["druck"].get()
            zeitwahl = settings["zeit"].get()
            verschiebung = verschiebung_mapping.get(zeitwahl, 0)
            ziel_blatt = druckwahl

            if ziel_blatt not in wb_ziel.sheetnames:
                wb_ziel.create_sheet(ziel_blatt)
            ws_ziel = wb_ziel[ziel_blatt]

            zeilennummern = [int(z.get()) if z.get().isdigit() else None for z in settings["zeilen"]]

            verschobene_zielzellen = []
            for z in zielzellen:
                col_letter = ''.join([c for c in z if c.isalpha()])
                row_number = ''.join([c for c in z if c.isdigit()])
                start_col_index = column_index_from_string(col_letter) + verschiebung
                neue_spalte = get_column_letter(start_col_index)
                verschobene_zielzellen.append(f"{neue_spalte}{row_number}")

            for i, blatt in enumerate(quellblaetter):
                zeile_x = zeilennummern[i]
                if zeile_x is None:
                    continue

                try:
                    wb_quelle = load_workbook(pfad, data_only=True, read_only=True)
                    if blatt not in wb_quelle.sheetnames:
                        continue
                    ws_quelle = wb_quelle[blatt]

                    # Nur die benötigte Zeile auslesen (ab Spalte B)
                    row_values = []
                    for row in ws_quelle.iter_rows(min_row=zeile_x, max_row=zeile_x, min_col=2):
                        for cell in row:
                            if cell.value is not None:
                                row_values.append(cell.value)
                    wb_quelle.close()

                    if not row_values:
                        continue

                    # Werte transponiert in Ziel schreiben (Zeile → Spalte)
                    ziel_zelle = verschobene_zielzellen[i]
                    col_letter = ''.join([c for c in ziel_zelle if c.isalpha()])
                    row_number = int(''.join([c for c in ziel_zelle if c.isdigit()]))
                    startcol = column_index_from_string(col_letter)
                    startrow = row_number

                    for idx, val in enumerate(row_values):
                        ws_ziel.cell(row=startrow + idx, column=startcol, value=val)

                except Exception as e:
                    messagebox.showwarning("Fehler", f"Fehler beim Lesen von {blatt} in {os.path.basename(pfad)}:\n{e}")
                    continue

        wb_ziel.save(ziel_datei)
        messagebox.showinfo("Fertig", "Daten erfolgreich übertragen (transponiert).")
    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler beim Schreiben:\n{e}")

# --- GUI ---
tk.Button(root, text="Quelldateien wählen", command=select_quellen).pack(pady=5)
tk.Button(root, text="Zieldatei wählen", command=select_ziel).pack(pady=5)
tk.Label(root, textvariable=ziel_var, fg="blue", wraplength=900).pack(pady=2)

canvas = tk.Canvas(root, height=500)
scrollbar = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
scroll_frame = tk.Frame(canvas)

scroll_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)
canvas.create_window((0,0), window=scroll_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

tk.Button(root, text="Daten übertragen", command=starten, bg="green", fg="white").pack(pady=10)

root.mainloop()
